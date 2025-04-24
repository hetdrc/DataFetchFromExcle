# Fetch multiple users from an Excel (.xlsx) file, attempt to log in with each set of credentials, and update the login status back into the Excel file.
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import load_workbook # type: ignore

driver = webdriver.Chrome()
driver.maximize_window()

# Excel file
workbook = load_workbook(filename="D:/Automation/MultipleLogin/MultipleLogin.xlsx")
sheet = workbook.active

for row in range(2, sheet.max_row + 1):
    username = sheet.cell(row=row, column=1).value
    password = sheet.cell(row=row, column=2).value
    print("Username:", username, " Password:", password)

    driver.get("https://uat.carematch.com/homecare/login")

    try:
        # Enter test 123 credentials
        if password and username:
            driver.find_element(By.XPATH, '//*[@id="EmailID"]').send_keys(username)
            driver.find_element(By.XPATH, '//*[@id="Password"]').send_keys(password)
            driver.find_element(By.XPATH,"/html/body/div[3]/section/div/div/div/form/div/div[3]/button").click()
        else:
            # print("Data missing")
            # break
            pass
        time.sleep(3)  # Wait for page to load

        if driver.current_url == "https://uat.carematch.com/homecare/dash-login":
            sheet.cell(row=row, column=3).value = "Success"
            print("Login successful")
        else:   
            sheet.cell(row=row, column=3).value = "Failed"
            print("Login failed")

    except Exception as e:
        print(f"An error occurred: {e}")

workbook.save("MultipleLogin.xlsx")

driver.quit()
